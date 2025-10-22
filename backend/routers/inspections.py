from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timedelta
import os
import uuid
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, KeepTogether
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.utils import ImageReader
from io import BytesIO
import base64
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import (
    User,
    Inspection,
    InspectionResponse,
    InspectionFile,
    Form,
    FormField,
    FieldType,
    InspectionStatus as ModelInspectionStatus,
    PassHoldStatus as ModelPassHoldStatus,
)
from schemas import (
    InspectionCreate,
    InspectionUpdate,
    InspectionResponse as InspectionResponseSchema,
    InspectionStatus as SchemaInspectionStatus,
    PassHoldStatus as SchemaPassHoldStatus,
)
from auth import get_current_user, require_role
from utils.flag_evaluator import evaluate_flag_conditions
from utils.logging_config import get_logger, log_file_upload_event

logger = get_logger(__name__)
router = APIRouter()

@router.get("/", response_model=List[InspectionResponseSchema])
async def get_inspections(
    skip: int = 0,
    limit: int = 100,
    status_filter: Optional[SchemaInspectionStatus] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get inspections based on user role"""
    query = db.query(Inspection).join(Form, Inspection.form_id == Form.id).join(User, Inspection.inspector_id == User.id)
    
    # Filter based on user role
    if current_user.role.value == "user":
        # Users can only see their own inspections
        query = query.filter(Inspection.inspector_id == current_user.id)
    elif current_user.role.value in ["supervisor", "management", "admin"]:
        # Supervisors, management, and admins can see all inspections
        pass
    
    if status_filter:
        status_filter_enum = ModelInspectionStatus(status_filter.value)
        query = query.filter(Inspection.status == status_filter_enum)
    
    inspections = query.offset(skip).limit(limit).all()
    
    # Add computed fields
    for inspection in inspections:
        inspection.has_flags = any(response.is_flagged for response in inspection.responses)
        # Get form name and inspector username
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        inspector = db.query(User).filter(User.id == inspection.inspector_id).first()
        inspection.form_name = form.form_name if form else None
        inspection.inspector_username = inspector.username if inspector else None
    
    return inspections

@router.get("/my-inspections", response_model=List[InspectionResponseSchema])
async def get_my_inspections(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get current user's inspections"""
    inspections = db.query(Inspection).filter(
        Inspection.inspector_id == current_user.id
    ).offset(skip).limit(limit).all()
    
    # Add has_flags computed field
    for inspection in inspections:
        inspection.has_flags = any(response.is_flagged for response in inspection.responses)
    
    return inspections

@router.get("/export-excel")
async def export_inspections_to_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    form_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export inspections to Excel with date filtering"""
    # Build query
    query = db.query(Inspection)
    
    # Filter based on user role
    if current_user.role.value == "user":
        query = query.filter(Inspection.inspector_id == current_user.id)
    
    # Apply filters
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Inspection.created_at >= start_dt)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid start_date format. Use YYYY-MM-DD"
            )
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            # Add one day to include the entire end date
            end_dt = end_dt + timedelta(days=1)
            query = query.filter(Inspection.created_at < end_dt)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid end_date format. Use YYYY-MM-DD"
            )
    
    if form_id:
        query = query.filter(Inspection.form_id == form_id)
    
    if status_filter:
        try:
            status_filter_enum = ModelInspectionStatus(status_filter)
            query = query.filter(Inspection.status == status_filter_enum)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid status_filter. Must be one of: {', '.join([s.value for s in ModelInspectionStatus])}"
            )
    
    inspections = query.all()
    
    if not inspections:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No inspections found with the specified filters"
        )
    
    # Create Enhanced Excel workbook with multiple sheets
    wb = Workbook()
    
    # Remove default sheet and create custom sheets
    wb.remove(wb.active)
    
    # Create Summary Sheet
    summary_ws = wb.create_sheet("📊 Summary")
    
    # Create Detailed Data Sheet
    detail_ws = wb.create_sheet("📋 Detailed Data")
    
    # Create Forms Overview Sheet
    forms_ws = wb.create_sheet("📝 Forms Overview")
    
    # Define enhanced styles
    # Header styles with gradient effect
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    header_fill = PatternFill(start_color="1E40AF", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Subheader styles
    subheader_font = Font(bold=True, color="1E40AF", size=11, name="Calibri")
    subheader_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    subheader_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data cell styles
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    
    # Status-specific styles
    pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    hold_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Enhanced borders
    thin_border = Border(
        left=Side(style='thin', color="D1D5DB"),
        right=Side(style='thin', color="D1D5DB"),
        top=Side(style='thin', color="D1D5DB"),
        bottom=Side(style='thin', color="D1D5DB")
    )
    
    thick_border = Border(
        left=Side(style='medium', color="1E40AF"),
        right=Side(style='medium', color="1E40AF"),
        top=Side(style='medium', color="1E40AF"),
        bottom=Side(style='medium', color="1E40AF")
    )
    
    # === POPULATE SUMMARY SHEET ===
    # Summary sheet title
    summary_ws['A1'] = "INSPECTION EXPORT SUMMARY"
    summary_ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1E40AF')
    summary_ws.merge_cells('A1:D1')
    summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Export details section
    summary_ws['A3'] = "Export Details"
    summary_ws['A3'].font = Font(name='Calibri', size=12, bold=True, color='1E40AF')
    summary_ws['A3'].fill = subheader_fill
    
    summary_ws['A4'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_ws['A5'] = f"Total Inspections: {len(inspections)}"
    summary_ws['A6'] = f"Date Range: {start_date or 'All'} to {end_date or 'All'}"
    
    # Get form name if form_id is provided
    form_name_display = "All Forms"
    if form_id:
        form = db.query(Form).filter(Form.id == form_id).first()
        if form:
            form_name_display = form.form_name
    
    summary_ws['A7'] = f"Form Filter: {form_name_display}"
    summary_ws['A8'] = f"Status Filter: {status_filter or 'All Statuses'}"
    
    # Statistics section
    total_inspections = len(inspections)
    status_counts = {}
    for inspection in inspections:
        status = inspection.status.value
        status_counts[status] = status_counts.get(status, 0) + 1
    
    summary_ws['A10'] = "Status Distribution"
    summary_ws['A10'].font = Font(name='Calibri', size=12, bold=True, color='1E40AF')
    summary_ws['A10'].fill = subheader_fill
    
    # Status distribution headers
    summary_ws['A11'] = "Status"
    summary_ws['B11'] = "Count"
    summary_ws['C11'] = "Percentage"
    
    for cell in ['A11', 'B11', 'C11']:
        summary_ws[cell].font = subheader_font
        summary_ws[cell].fill = subheader_fill
        summary_ws[cell].alignment = subheader_alignment
        summary_ws[cell].border = thin_border
    
    # Status data rows
    row = 12
    for status, count in status_counts.items():
        percentage = f"{(count/total_inspections*100):.1f}%" if total_inspections > 0 else "0%"
        
        summary_ws[f'A{row}'] = status.upper()
        summary_ws[f'B{row}'] = count
        summary_ws[f'C{row}'] = percentage
        
        # Apply status-specific styling
        if status == 'accepted':
            fill = pass_fill
        elif status == 'rejected':
            fill = hold_fill
        else:
            fill = pending_fill
        
        for col in ['A', 'B', 'C']:
            cell = summary_ws[f'{col}{row}']
            cell.fill = fill
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        row += 1
    
    # Form distribution section
    form_counts = {}
    for inspection in inspections:
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        form_name = form.form_name if form else "Unknown Form"
        form_counts[form_name] = form_counts.get(form_name, 0) + 1
    
    summary_ws[f'A{row + 1}'] = "📝 Form Distribution"
    summary_ws[f'A{row + 1}'].font = Font(name='Calibri', size=12, bold=True, color='1E40AF')
    summary_ws[f'A{row + 1}'].fill = subheader_fill
    
    # Form distribution headers
    row += 2
    summary_ws[f'A{row}'] = "Form Name"
    summary_ws[f'B{row}'] = "Count"
    summary_ws[f'C{row}'] = "Percentage"
    
    for cell in [f'A{row}', f'B{row}', f'C{row}']:
        summary_ws[cell].font = subheader_font
        summary_ws[cell].fill = subheader_fill
        summary_ws[cell].alignment = subheader_alignment
        summary_ws[cell].border = thin_border
    
    # Form data rows
    row += 1
    for form_name, count in form_counts.items():
        percentage = f"{(count/total_inspections*100):.1f}%" if total_inspections > 0 else "0%"
        
        summary_ws[f'A{row}'] = form_name
        summary_ws[f'B{row}'] = count
        summary_ws[f'C{row}'] = percentage
        
        for col in ['A', 'B', 'C']:
            cell = summary_ws[f'{col}{row}']
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        row += 1
    
    # Auto-adjust column widths for summary sheet
    for column in summary_ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            # Get column letter from the first non-merged cell
            if column_letter is None and hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # === POPULATE FORMS OVERVIEW SHEET ===
    # Forms overview title
    forms_ws['A1'] = "📝 FORMS OVERVIEW"
    forms_ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1E40AF')
    forms_ws.merge_cells('A1:E1')
    forms_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Get all unique forms from inspections
    unique_forms = {}
    for inspection in inspections:
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        if form and form.id not in unique_forms:
            unique_forms[form.id] = form
    
    # Forms overview headers
    forms_ws['A3'] = "Form Name"
    forms_ws['B3'] = "Total Fields"
    forms_ws['C3'] = "Field Types"
    forms_ws['D3'] = "Inspections Count"
    forms_ws['E3'] = "Last Used"
    
    for cell in ['A3', 'B3', 'C3', 'D3', 'E3']:
        forms_ws[cell].font = subheader_font
        forms_ws[cell].fill = subheader_fill
        forms_ws[cell].alignment = subheader_alignment
        forms_ws[cell].border = thin_border
    
    # Forms data rows
    row = 4
    for form_id, form in unique_forms.items():
        # Get form fields
        form_fields = db.query(FormField).filter(FormField.form_id == form_id).all()
        field_types = list(set([field.field_type for field in form_fields]))
        
        # Get inspection count for this form
        inspection_count = len([i for i in inspections if i.form_id == form_id])
        
        # Get last used date
        last_inspection = max([i for i in inspections if i.form_id == form_id], 
                            key=lambda x: x.created_at, default=None)
        last_used = last_inspection.created_at.strftime('%Y-%m-%d') if last_inspection else "Never"
        
        forms_ws[f'A{row}'] = form.form_name
        forms_ws[f'B{row}'] = len(form_fields)
        forms_ws[f'C{row}'] = ", ".join([field_type.value for field_type in field_types])
        forms_ws[f'D{row}'] = inspection_count
        forms_ws[f'E{row}'] = last_used
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = forms_ws[f'{col}{row}']
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        row += 1
    
    # Auto-adjust column widths for forms overview sheet
    for column in forms_ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            # Get column letter from the first non-merged cell
            if column_letter is None and hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            forms_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Get all unique form fields from the inspections
    form_fields_map = {}
    for inspection in inspections:
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        if form and form.id not in form_fields_map:
            form_fields_map[form.id] = {
                'name': form.form_name,
                'fields': sorted(form.fields, key=lambda f: f.field_order)
            }
    
    # Create header row
    headers = [
        "Inspection ID",
        "Form Name",
        "Inspector",
        "Status",
        "Created Date",
        "Updated Date",
        "Reviewed By",
        "Reviewed Date",
        "Rejection Reason"
    ]
    
    # Add dynamic field headers (collect all unique fields)
    all_fields = []
    field_ids_seen = set()
    for form_data in form_fields_map.values():
        for field in form_data['fields']:
            if field.id not in field_ids_seen:
                all_fields.append(field)
                field_ids_seen.add(field.id)
                headers.append(f"{field.field_name} ({field.field_type})")
    
    # === DETAILED DATA SHEET ===
    ws = detail_ws
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    row_num = 2
    for inspection in inspections:
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        inspector = db.query(User).filter(User.id == inspection.inspector_id).first()
        reviewer = None
        if inspection.reviewed_by:
            reviewer = db.query(User).filter(User.id == inspection.reviewed_by).first()
        
        # Basic inspection info
        row_data = [
            inspection.id,
            form.form_name if form else "N/A",
            inspector.username if inspector else "N/A",
            inspection.status.value.upper(),
            inspection.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            inspection.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            reviewer.username if reviewer else "",
            inspection.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if inspection.reviewed_at else "",
            inspection.rejection_reason or ""
        ]
        
        # Get responses mapped by field_id
        responses_map = {resp.field_id: resp for resp in inspection.responses}
        
        # Add field responses
        for field in all_fields:
            field_response = responses_map.get(field.id)
            cell_value = ""
            
            if field_response:
                if field_response.response_value:
                    if field.field_type == 'signature':
                        cell_value = "[Digital Signature]"
                    elif field.field_type == 'photo':
                        cell_value = f"[Photo: {field_response.response_value}]"
                    else:
                        cell_value = str(field_response.response_value)
                
                if field_response.measurement_value is not None:
                    cell_value = str(field_response.measurement_value)
                    if field.field_options and 'unit' in field.field_options:
                        cell_value += f" {field.field_options['unit']}"
                
                if field_response.pass_hold_status:
                    # Handle both enum and string types for pass_hold_status
                    if hasattr(field_response.pass_hold_status, 'value'):
                        status_text = field_response.pass_hold_status.value.upper()
                    else:
                        status_text = str(field_response.pass_hold_status).upper()
                    cell_value += f" [{status_text}]" if cell_value else f"[{status_text}]"
            else:
                cell_value = "—"
            
            row_data.append(cell_value)
        
        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = cell_font
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory buffer
    excel_filename = f"inspections_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_buffer = BytesIO()
    
    wb.save(excel_buffer)
    
    # Get Excel content from buffer
    excel_buffer.seek(0)
    excel_content = excel_buffer.getvalue()
    excel_buffer.close()
    
    # Return the Excel file directly from memory
    from fastapi.responses import Response
    return Response(
        content=excel_content,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={excel_filename}"}
    )

@router.get("/{inspection_id}", response_model=InspectionResponseSchema)
async def get_inspection(
    inspection_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get inspection by ID"""
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    if (current_user.role.value == "user" and 
        inspection.inspector_id != current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # Add has_flags computed field
    inspection.has_flags = any(response.is_flagged for response in inspection.responses)
    
    return inspection

@router.post("/", response_model=InspectionResponseSchema)
async def create_inspection(
    inspection: InspectionCreate,
    current_user: User = Depends(require_role(["user", "admin"])),
    db: Session = Depends(get_db)
):
    """Create new inspection"""
    # Verify form exists
    form = db.query(Form).filter(Form.id == inspection.form_id).first()
    if not form:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found"
        )
    
    # Create inspection
    db_inspection = Inspection(
        form_id=inspection.form_id,
        inspector_id=current_user.id,
        status=ModelInspectionStatus.draft
    )
    
    db.add(db_inspection)
    db.commit()
    db.refresh(db_inspection)
    
    # Create responses
    for response_data in inspection.responses:
        # Handle conditional fields (field_id can be None)
        if response_data.field_id is None:
            logger.debug(f"Saving conditional field response: {response_data.response_value}")
        else:
            logger.debug(f"Saving field response for field_id {response_data.field_id}: {response_data.response_value}")
            
        # Normalize pass_hold_status to raw string value
        pass_hold_status = None
        if response_data.pass_hold_status is not None:
            if isinstance(response_data.pass_hold_status, SchemaPassHoldStatus):
                raw_value = response_data.pass_hold_status.value
            elif isinstance(response_data.pass_hold_status, ModelPassHoldStatus):
                raw_value = response_data.pass_hold_status.value
            else:
                raw_value = str(response_data.pass_hold_status)
            
            # Ensure we store the actual string value
            if raw_value in ["pass", "hold"]:
                pass_hold_status = raw_value

        # Evaluate flag conditions for this response
        is_flagged = False
        if response_data.field_id is not None:
            # Get the field to check for flag conditions
            field = db.query(FormField).filter(FormField.id == response_data.field_id).first()
            if field and field.flag_conditions:
                is_flagged = evaluate_flag_conditions(
                    field.flag_conditions,
                    field.field_type,
                    response_data.response_value,
                    response_data.measurement_value,
                    pass_hold_status
                )

        db_response = InspectionResponse(
            inspection_id=db_inspection.id,
            field_id=response_data.field_id,
            response_value=response_data.response_value,
            measurement_value=response_data.measurement_value,
            pass_hold_status=pass_hold_status,
            is_flagged=is_flagged
        )
        db.add(db_response)
    
    db.commit()
    db.refresh(db_inspection)
    
    return db_inspection

@router.put("/{inspection_id}", response_model=InspectionResponseSchema)
async def update_inspection(
    inspection_id: int,
    inspection_update: InspectionUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update inspection"""
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    can_edit = False
    if current_user.role.value == "admin":
        can_edit = True
    elif current_user.role.value == "user" and inspection.inspector_id == current_user.id:
        # Users can only edit their own draft inspections
        can_edit = inspection.status == ModelInspectionStatus.draft
    elif current_user.role.value == "supervisor":
        # Supervisors can review submitted inspections
        can_edit = inspection.status == ModelInspectionStatus.submitted
    
    if not can_edit:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions or inspection cannot be modified"
        )
    
    # Update inspection
    update_data = inspection_update.dict(exclude_unset=True)

    # Handle responses update
    responses_data = update_data.pop("responses", None)
    if responses_data is not None:
        # Delete existing responses
        db.query(InspectionResponse).filter(
            InspectionResponse.inspection_id == inspection_id
        ).delete()
        
        # Create new responses with flag evaluation
        for response_data in responses_data:
            # Convert Pydantic model to dict if needed
            if hasattr(response_data, 'dict'):
                response_dict = response_data.dict()
            else:
                response_dict = response_data
            
            # Normalize pass_hold_status to raw string value
            pass_hold_status = None
            pass_hold_value = response_dict.get('pass_hold_status')
            if pass_hold_value is not None:
                if isinstance(pass_hold_value, SchemaPassHoldStatus):
                    raw_value = pass_hold_value.value
                elif isinstance(pass_hold_value, ModelPassHoldStatus):
                    raw_value = pass_hold_value.value
                else:
                    raw_value = str(pass_hold_value)
                
                # Ensure we store the actual string value
                if raw_value in ["pass", "hold"]:
                    pass_hold_status = raw_value

            # Evaluate flag conditions for this response
            is_flagged = False
            field_id = response_dict.get('field_id')
            if field_id is not None:
                # Get the field to check for flag conditions
                field = db.query(FormField).filter(FormField.id == field_id).first()
                if field and field.flag_conditions:
                    is_flagged = evaluate_flag_conditions(
                        field.flag_conditions,
                        field.field_type,
                        response_dict.get('response_value'),
                        response_dict.get('measurement_value'),
                        pass_hold_status
                    )

            db_response = InspectionResponse(
                inspection_id=inspection_id,
                field_id=field_id,
                response_value=response_dict.get('response_value'),
                measurement_value=response_dict.get('measurement_value'),
                pass_hold_status=pass_hold_status,
                is_flagged=is_flagged
            )
            db.add(db_response)

    status_value = update_data.pop("status", None)
    status_enum = None
    if status_value is not None:
        status_enum = ModelInspectionStatus(
            status_value.value if isinstance(status_value, SchemaInspectionStatus) else status_value
        )
        inspection.status = status_enum

    for field, value in update_data.items():
        setattr(inspection, field, value)

    if status_enum in (ModelInspectionStatus.accepted, ModelInspectionStatus.rejected):
        inspection.reviewed_by = current_user.id
        inspection.reviewed_at = datetime.utcnow()
    
    db.commit()
    db.refresh(inspection)
    
    return inspection

@router.post("/{inspection_id}/submit")
async def submit_inspection(
    inspection_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Submit inspection for review"""
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    if (current_user.role.value == "user" and 
        inspection.inspector_id != current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    if inspection.status != ModelInspectionStatus.draft:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only draft inspections can be submitted"
        )

    inspection.status = ModelInspectionStatus.submitted
    db.commit()
    
    return {"message": "Inspection submitted successfully"}

@router.delete("/{inspection_id}")
async def delete_inspection(
    inspection_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete inspection"""
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    can_delete = False
    if current_user.role.value == "admin":
        can_delete = True
    elif (current_user.role.value == "user" and 
          inspection.inspector_id == current_user.id and 
          inspection.status == ModelInspectionStatus.draft):
        can_delete = True
    
    if not can_delete:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    try:
        # Delete related data first to avoid foreign key constraint errors
        # Delete inspection responses
        db.query(InspectionResponse).filter(InspectionResponse.inspection_id == inspection_id).delete()
        
        # Delete inspection files
        db.query(InspectionFile).filter(InspectionFile.inspection_id == inspection_id).delete()
        
        # Now delete the inspection itself
        db.delete(inspection)
        db.commit()
        
        return {"message": "Inspection deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting inspection: {str(e)}"
        )

@router.post("/{inspection_id}/upload-file")
async def upload_file(
    inspection_id: int,
    field_id: int,
    file_type: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upload file for inspection field with security validation"""
    # Import file validation utility
    from utils.file_validation import FileValidator
    
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    if (current_user.role.value == "user" and 
        inspection.inspector_id != current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # SECURITY: Validate file before processing
    # Define allowed file types based on file_type parameter
    allowed_types = []
    if file_type in ['image', 'photo']:
        allowed_types = ['image']
    elif file_type in ['document', 'report']:
        allowed_types = ['document']
    elif file_type in ['spreadsheet', 'data']:
        allowed_types = ['spreadsheet']
    else:
        # Default to images and documents for inspection files
        allowed_types = ['image', 'document']
    
    # Validate file with security checks
    try:
        file_metadata = await FileValidator.validate_upload_file(
            file=file,
            allowed_types=allowed_types,
            max_size=10 * 1024 * 1024  # 10MB limit
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File validation failed: {str(e)}"
        )
    
    # Create uploads directory if it doesn't exist
    upload_dir = "uploads"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate unique filename using sanitized name
    file_extension = os.path.splitext(file_metadata['safe_filename'])[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Save file securely
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to save file"
        )
    
    # Save file info to database with security metadata
    db_file = InspectionFile(
        inspection_id=inspection_id,
        field_id=field_id,
        file_name=file_metadata['safe_filename'],  # Use sanitized filename
        file_path=file_path,
        file_type=file_type
    )
    
    db.add(db_file)
    db.commit()
    
    return {
        "message": "File uploaded successfully", 
        "file_id": db_file.id,
        "original_filename": file_metadata['original_filename'],
        "safe_filename": file_metadata['safe_filename'],
        "file_hash": file_metadata['hash']
    }

def process_image_for_pdf(base64_data: str, max_width: float, max_height: float, field_id: int, image_type: str = "image"):
    """
    Robust image processing function for PDF export
    Returns ReportLab Image object or None if processing fails
    """
    try:
        logger.info(f"Processing {image_type} for field {field_id}: Starting processing")
        
        # Validate input
        if not base64_data or not isinstance(base64_data, str):
            logger.error(f"Processing {image_type} for field {field_id}: Invalid base64 data")
            return None
        
        # Clean base64 data
        clean_data = base64_data.strip()
        
        # Remove data URL prefix if present
        if clean_data.startswith('data:image'):
            if ',' not in clean_data:
                logger.error(f"Processing {image_type} for field {field_id}: Invalid data URL format")
                return None
            clean_data = clean_data.split(',')[1]
        
        # Add padding if necessary
        missing_padding = len(clean_data) % 4
        if missing_padding:
            clean_data += '=' * (4 - missing_padding)
        
        # Decode base64
        try:
            image_data = base64.b64decode(clean_data)
        except Exception as e:
            logger.error(f"Processing {image_type} for field {field_id}: Base64 decode error: {e}")
            return None
        
        # Validate decoded data
        if len(image_data) < 100:  # Minimum reasonable image size
            logger.error(f"Processing {image_type} for field {field_id}: Image data too small")
            return None
        
        # Create PIL Image
        image_buffer = BytesIO(image_data)
        try:
            pil_image = PILImage.open(image_buffer)
            pil_image.verify()  # Verify image integrity
            
            # Reopen image after verify (verify closes the image)
            image_buffer.seek(0)
            pil_image = PILImage.open(image_buffer)
        except Exception as e:
            logger.error(f"Processing {image_type} for field {field_id}: PIL Image error: {e}")
            return None
        
        # Convert to RGB if necessary
        if pil_image.mode not in ['RGB', 'L']:
            pil_image = pil_image.convert('RGB')
        
        # Calculate optimal size while maintaining aspect ratio
        original_width, original_height = pil_image.size
        aspect_ratio = original_width / original_height
        
        # Calculate new dimensions
        if aspect_ratio > max_width / max_height:
            # Width is the limiting factor
            new_width = max_width
            new_height = max_width / aspect_ratio
        else:
            # Height is the limiting factor
            new_height = max_height
            new_width = max_height * aspect_ratio
        
        # Ensure minimum size
        min_size = 20  # Minimum 20 points
        new_width = max(new_width, min_size)
        new_height = max(new_height, min_size)
        
        logger.info(f"Processing {image_type} for field {field_id}: Calculated size: {new_width}x{new_height}")
        
        # Resize image for better quality
        resize_factor = 2  # Higher resolution for better quality
        pixel_width = int(new_width * resize_factor)
        pixel_height = int(new_height * resize_factor)
        
        resized_image = pil_image.resize((pixel_width, pixel_height), PILImage.Resampling.LANCZOS)
        
        # Save to buffer with high quality
        final_buffer = BytesIO()
        if image_type == "signature":
            # Use PNG for signatures to preserve transparency
            resized_image.save(final_buffer, format='PNG', optimize=True)
        else:
            # Use JPEG for photos with high quality
            resized_image.save(final_buffer, format='JPEG', quality=90, optimize=True)
        
        # Create ReportLab Image using buffer directly
        final_buffer.seek(0)
        reportlab_image = Image(final_buffer, width=new_width, height=new_height)
        
        logger.info(f"Processing {image_type} for field {field_id}: Successfully created ReportLab Image")
        return reportlab_image
        
    except Exception as e:
        logger.error(f"Processing {image_type} for field {field_id}: Unexpected error: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None

@router.get("/{inspection_id}/export-pdf")
async def export_inspection_to_pdf(
    inspection_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export inspection to PDF with questions on left and answers on right"""
    # Get inspection with all related data
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inspection not found"
        )
    
    # Check permissions
    if (current_user.role.value == "user" and 
        inspection.inspector_id != current_user.id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # Get form and fields
    form = db.query(Form).filter(Form.id == inspection.form_id).first()
    if not form:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found"
        )
    
    # Get inspector info
    inspector = db.query(User).filter(User.id == inspection.inspector_id).first()
    
    # Create PDF in memory buffer
    pdf_filename = f"inspection_{inspection_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_buffer = BytesIO()
    
    # Create the PDF document using memory buffer
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, 
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=30)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define enhanced styles
    styles = getSampleStyleSheet()
    
    # Enhanced title style with gradient-like effect
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=16,
        spaceBefore=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Enhanced heading style
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=8,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderPadding=0,
        leftIndent=0
    )
    
    # Enhanced section heading style
    section_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading3'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#f3f4f6'),
        borderWidth=1,
        borderColor=colors.HexColor('#d1d5db'),
        borderPadding=6,
        leftIndent=6
    )
    
    # Enhanced normal style
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4,
        leading=12
    )
    
    # Status indicator styles
    pass_style = ParagraphStyle(
        'PassStatus',
        parent=normal_style,
        textColor=colors.HexColor('#059669'),
        fontName='Helvetica-Bold'
    )
    
    hold_style = ParagraphStyle(
        'HoldStatus',
        parent=normal_style,
        textColor=colors.HexColor('#dc2626'),
        fontName='Helvetica-Bold'
    )
    
    # Enhanced title with company branding
    title = Paragraph(f"<b>INSPECTION REPORT</b>", title_style)
    elements.append(title)
    
    # Subtitle with form name
    subtitle = Paragraph(f"<i>{form.form_name}</i>", ParagraphStyle(
        'Subtitle',
        parent=normal_style,
        fontSize=12,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER,
        spaceAfter=20
    ))
    elements.append(subtitle)
    elements.append(Spacer(1, 16))
    
    # Enhanced Inspection Info Section
    info_section = Paragraph("<b>INSPECTION DETAILS</b>", section_style)
    elements.append(info_section)
    elements.append(Spacer(1, 8))
    
    # Status indicator with black text for better visibility
    status_color = colors.black
    
    status_indicator = f"{inspection.status.value.upper()}"
    
    # Calculate flag summary for inspection details
    flag_summary = {"flagged_count": 0, "total_responses": 0, "pass_count": 0, "hold_count": 0}
    responses_map = {resp.field_id: resp for resp in inspection.responses}
    
    # Import flag evaluator
    from utils.flag_evaluator import FlagEvaluator
    
    for field in form.fields:
        field_response = responses_map.get(field.id)
        if field_response:
            flag_summary["total_responses"] += 1
            
            # Check if flagged
            if field.flag_conditions:
                is_flagged = FlagEvaluator.evaluate_field_response(
                    field_type=field.field_type,
                    response_value=field_response.response_value,
                    measurement_value=field_response.measurement_value,
                    flag_conditions=field.flag_conditions
                )
                if is_flagged:
                    flag_summary["flagged_count"] += 1
            
            # Count pass/hold status
            if field_response.pass_hold_status:
                status_value = field_response.pass_hold_status.value if hasattr(field_response.pass_hold_status, 'value') else str(field_response.pass_hold_status)
                if status_value.upper() == 'PASS':
                    flag_summary["pass_count"] += 1
                elif status_value.upper() == 'HOLD':
                    flag_summary["hold_count"] += 1
    
    # Determine overall flag status
    overall_flag = "FLAGGED" if flag_summary["flagged_count"] > 0 else "CLEAR"
    flag_color = colors.HexColor('#dc2626') if flag_summary["flagged_count"] > 0 else colors.HexColor('#059669')
    
    info_data = [
        ['Inspection ID:', str(inspection.id)],
        ['Form:', form.form_name],
        ['Inspector:', inspector.username if inspector else 'N/A'],
        ['Flag Status:', overall_flag],
        ['Status:', status_indicator],
        ['Created:', inspection.created_at.strftime('%Y-%m-%d %H:%M:%S')],
        ['Updated:', inspection.updated_at.strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    if inspection.reviewed_by:
        reviewer = db.query(User).filter(User.id == inspection.reviewed_by).first()
        info_data.append(['Reviewed By:', reviewer.username if reviewer else 'N/A'])
        if inspection.reviewed_at:
            info_data.append(['Reviewed At:', inspection.reviewed_at.strftime('%Y-%m-%d %H:%M:%S')])
    
    if inspection.rejection_reason:
        info_data.append(['Rejection Reason:', inspection.rejection_reason])
    
    info_table = Table(info_data, colWidths=[2.2*inch, 3.8*inch])
    info_table.setStyle(TableStyle([
        # Header column styling
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (0, -1), 10),
        
        # Data column styling
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 0), (1, -1), 10),
        
        # Flag status row special styling
        ('BACKGROUND', (1, 3), (1, 3), flag_color),
        ('TEXTCOLOR', (1, 3), (1, 3), colors.white),
        ('FONTNAME', (1, 3), (1, 3), 'Helvetica-Bold'),
        
        # Status row special styling
        ('BACKGROUND', (1, 4), (1, 4), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (1, 4), (1, 4), colors.black),
        ('FONTNAME', (1, 4), (1, 4), 'Helvetica-Bold'),
        
        # General styling
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        
        # Rounded corners effect with alternating backgrounds
        *[('BACKGROUND', (1, i), (1, i), colors.HexColor('#ffffff') if i % 2 == 0 else colors.HexColor('#f8fafc'))
          for i in range(len(info_data)) if i != 3]  # Skip status row
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 24))
    
    # Enhanced Responses Section
    responses_title = Paragraph("<b>INSPECTION RESPONSES</b>", section_style)
    elements.append(responses_title)
    elements.append(Spacer(1, 12))
    
    # Get all responses mapped by field_id
    responses_map = {resp.field_id: resp for resp in inspection.responses}
    
    # Sort fields by field_order
    sorted_fields = sorted(form.fields, key=lambda f: f.field_order)
    
    # Group fields by type for better organization
    field_groups = {}
    for field in sorted_fields:
        field_type = field.field_type
        if field_type not in field_groups:
            field_groups[field_type] = []
        field_groups[field_type].append(field)
    
    # Field type icons mapping
    field_icons = {
        'text': '',
        'dropdown': '',
        'search_dropdown': '',
        'button': '',
        'photo': '',
        'signature': '',
        'measurement': '',
        'notes': '',
        'datetime': '',
        'time': ''
    }
    
    # Import flag evaluator
    from utils.flag_evaluator import FlagEvaluator
    
    # Create responses table with enhanced formatting
    response_data = [
        [
            Paragraph("<b>QUESTION</b>", ParagraphStyle('HeaderLeft', parent=normal_style, fontSize=11, textColor=colors.white, fontName='Helvetica-Bold')),
            Paragraph("<b>RESPONSE</b>", ParagraphStyle('HeaderRight', parent=normal_style, fontSize=11, textColor=colors.white, fontName='Helvetica-Bold'))
        ]
    ]
    
    # Track flagged rows for styling
    flagged_rows = []
    
    for field in sorted_fields:
        field_response = responses_map.get(field.id)
        logger.info(f"Processing field {field.id}: {field.field_name} ({field.field_type})")
        
        # Simplified question formatting
        question_text = field.field_name
        if field.is_required:
            question_text += " <font color='red'>*</font>"
        
        # Enhanced answer formatting with status indicators
        answer_text = ""
        answer_style = normal_style
        is_flagged = False
        
        if field_response:
            # Evaluate flag conditions for this response
            if field.flag_conditions:
                is_flagged = FlagEvaluator.evaluate_field_response(
                    field_type=field.field_type,
                    response_value=field_response.response_value,
                    measurement_value=field_response.measurement_value,
                    flag_conditions=field.flag_conditions
                )
            
            if field_response.response_value:
                # Handle different field types with appropriate formatting
                if field.field_type == 'signature' and field_response.response_value.startswith('data:image'):
                    # Process signature image using robust function with very small dimensions
                    signature_image = process_image_for_pdf(
                        field_response.response_value, 
                        max_width=1*inch, 
                        max_height=0.5*inch, 
                        field_id=field.id, 
                        image_type="signature"
                    )
                    
                    if signature_image:
                        # Create a very compact signature display
                        signature_content = KeepTogether([
                            Paragraph("<font color='#059669' size='6'>[Digital Signature]</font>", 
                                    ParagraphStyle('Answer', parent=normal_style, fontSize=6, spaceAfter=0, spaceBefore=0)),
                            signature_image
                        ])
                        
                        # Add this row with special handling for image
                        question_paragraph = Paragraph(
                            f"<b>{question_text}</b>",
                            ParagraphStyle('Question', parent=normal_style, fontSize=10, spaceAfter=2)
                        )
                        
                        response_data.append([question_paragraph, signature_content])
                        
                        # Track flagged rows if needed
                        if is_flagged:
                            flagged_rows.append(len(response_data) - 1)
                        
                        continue  # Skip the normal processing for this field
                    else:
                        answer_text = "<font color='#dc2626'>[Error processing signature]</font>"
                elif field.field_type == FieldType.photo:
                    if field_response.response_value:
                        # Use proper ReportLab units to prevent layout errors
                        photo_image = process_image_for_pdf(
                            field_response.response_value, 
                            max_width=1.5*inch, 
                            max_height=1*inch, 
                            field_id=field.id, 
                            image_type="photo"
                        )
                        if photo_image:
                            # Create a combined answer with text and image (similar to signature handling)
                            answer_paragraph = [
                                Paragraph("📷 <font color='#059669'>[Photo Evidence]</font>", 
                                        ParagraphStyle('Answer', parent=normal_style, fontSize=10, spaceAfter=2)),
                                photo_image
                            ]
                            
                            # Add this row with special handling for image
                            question_paragraph = Paragraph(
                                f"<b>{question_text}</b>",
                                ParagraphStyle('Question', parent=normal_style, fontSize=10, spaceAfter=2)
                            )
                            
                            response_data.append([question_paragraph, answer_paragraph])
                            
                            # Track flagged rows if needed
                            if is_flagged:
                                flagged_rows.append(len(response_data) - 1)
                            
                            continue  # Skip the normal processing for this field
                        else:
                            answer_text = "<font color='#dc2626'>[Error processing photo]</font>"
                    else:
                        answer_text = "📷 <font color='#dc2626'>[No photo provided]</font>"
                elif field.field_type == 'datetime':
                    answer_text = f"{field_response.response_value}"
                elif field.field_type == 'time':
                    answer_text = f"{field_response.response_value}"
                elif field.field_type == 'dropdown' or field.field_type == 'search_dropdown':
                    answer_text = f"{field_response.response_value}"
                else:
                    answer_text = str(field_response.response_value)
            
            # Handle measurement values with units
            if field_response.measurement_value is not None:
                unit = ""
                if field.field_options and 'unit' in field.field_options:
                    unit = f" {field.field_options['unit']}"
                answer_text = f"{field_response.measurement_value}{unit}"
            
            # Handle status indicators differently for KeepTogether vs text
            status_indicators = []
            
            # Enhanced pass/hold status with color coding
            if field_response.pass_hold_status:
                # Handle both enum and string types for pass_hold_status
                if hasattr(field_response.pass_hold_status, 'value'):
                    status_value = field_response.pass_hold_status.value.upper()
                else:
                    status_value = str(field_response.pass_hold_status).upper()
                if status_value == 'PASS':
                    status_indicator = f"<font color='#059669'><b>[{status_value}]</b></font>"
                else:  # HOLD
                    status_indicator = f"<font color='#dc2626'><b>[{status_value}]</b></font>"
                status_indicators.append(status_indicator)
            
            # Add flag indicator if flagged
            if is_flagged:
                status_indicators.append("<font color='#dc2626'><b>[FLAGGED]</b></font>")
            
            # Apply status indicators based on answer_text type
            if hasattr(answer_text, '__class__') and 'Image' in str(answer_text.__class__):
                # For Image objects (photos), create a simple list with image and status
                if status_indicators:
                    status_text = " ".join(status_indicators)
                    status_paragraph = Paragraph(status_text, ParagraphStyle('Status', parent=normal_style, fontSize=10, spaceAfter=2))
                    answer_text = [answer_text, status_paragraph]
                # If no status indicators, keep just the image
            else:
                # For text answers, append status indicators directly
                if status_indicators:
                    status_text = " " + " ".join(status_indicators)
                    answer_text += status_text if answer_text else status_text[1:]  # Remove leading space if no answer
        else:
            answer_text = "<font color='#9ca3af'>— No response provided —</font>"
        
        # Create formatted paragraphs
        question_paragraph = Paragraph(
            f"<b>{question_text}</b>",
            ParagraphStyle('Question', parent=normal_style, fontSize=10, spaceAfter=2)
        )
        
        # Handle different answer types (text vs list/Image for images)
        if isinstance(answer_text, (list, KeepTogether)) or (hasattr(answer_text, '__class__') and 'Image' in str(answer_text.__class__)):
            # For images (photos/signatures), answer_text is already a flowable
            answer_paragraph = answer_text
        else:
            # For text answers, create a Paragraph
            answer_paragraph = Paragraph(
                answer_text,
                ParagraphStyle('Answer', parent=normal_style, fontSize=10, spaceAfter=2)
            )
        
        response_data.append([question_paragraph, answer_paragraph])
        
        # Track flagged rows (add 1 because header is row 0)
        if is_flagged:
            flagged_rows.append(len(response_data) - 1)
    
    # Create the enhanced table with better row height control
    response_table = Table(response_data, colWidths=[3.2*inch, 3.3*inch], repeatRows=1)
    
    # Add row height constraints for signature rows to prevent layout errors
    signature_row_styles = []
    for i, row in enumerate(response_data):
        if len(row) > 1 and hasattr(row[1], '__class__') and 'KeepTogether' in str(row[1].__class__):
            # This is likely a signature row, limit its height
            signature_row_styles.append(('ROWBACKGROUNDS', (0, i), (-1, i), [colors.white]))
            signature_row_styles.append(('VALIGN', (0, i), (-1, i), 'TOP'))
    
    response_table.setStyle(TableStyle([
        # Header row styling with gradient-like effect
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Question column styling
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (0, -1), 10),
        
        # Answer column styling with alternating colors
        ('TEXTCOLOR', (1, 1), (1, -1), colors.HexColor('#374151')),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 1), (1, -1), 10),
        
        # Enhanced grid and spacing
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Alternating row colors for better readability
        *[('BACKGROUND', (1, i), (1, i), colors.HexColor('#ffffff') if i % 2 == 1 else colors.HexColor('#f9fafb'))
          for i in range(1, len(response_data)) if i not in flagged_rows],
        
        # Special styling for rows with PASS/HOLD status (non-flagged rows only)
        *[('BACKGROUND', (1, i), (1, i), colors.HexColor('#f0fdf4') if 'PASS' in str(response_data[i][1]) 
           else colors.HexColor('#fef2f2') if 'HOLD' in str(response_data[i][1]) 
           else colors.HexColor('#ffffff') if i % 2 == 1 else colors.HexColor('#f9fafb'))
          for i in range(1, len(response_data)) if i not in flagged_rows],
        
        # Red background for flagged abnormal data
        *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fee2e2'))  # Light red background
          for i in flagged_rows],
        *[('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#991b1b'))  # Dark red text
          for i in flagged_rows],
        *[('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold')  # Bold text for flagged rows
          for i in flagged_rows],
        
        # Apply signature row constraints
        *signature_row_styles
    ]))
    
    elements.append(response_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF content from buffer
    pdf_buffer.seek(0)
    pdf_content = pdf_buffer.getvalue()
    pdf_buffer.close()
    
    # Return the PDF file directly from memory
    from fastapi.responses import Response
    return Response(
        content=pdf_content,
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment; filename={pdf_filename}"}
    )
