#!/usr/bin/env python3
"""
Test script to generate actual PDF and Excel files to examine formatting
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from sqlalchemy.orm import Session
from database import get_db
from models import Inspection, User, Form, InspectionResponse, FormField
from datetime import datetime
import tempfile

# Import the actual export functions
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_sample_pdf():
    """Generate a sample PDF to test formatting"""
    print("Generating sample PDF...")
    
    db = next(get_db())
    
    try:
        # Get a sample inspection with responses
        inspection = db.query(Inspection).filter(
            Inspection.id.in_(
                db.query(InspectionResponse.inspection_id).distinct()
            )
        ).first()
        
        if not inspection:
            print("❌ No inspection with responses found")
            return False
            
        print(f"✓ Using inspection ID: {inspection.id}")
        
        # Get related data
        form = db.query(Form).filter(Form.id == inspection.form_id).first()
        inspector = db.query(User).filter(User.id == inspection.inspector_id).first()
        responses = db.query(InspectionResponse).filter(
            InspectionResponse.inspection_id == inspection.id
        ).all()
        
        # Create PDF
        pdf_filename = f"sample_inspection_{inspection.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf_path = os.path.join("uploads", pdf_filename)
        
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1E40AF')
        )
        story.append(Paragraph("INSPECTION REPORT", title_style))
        story.append(Spacer(1, 20))
        
        # Inspection Info Table
        info_data = [
            ['Inspection ID:', str(inspection.id)],
            ['Form Name:', form.form_name if form else 'N/A'],
            ['Inspector:', inspector.username if inspector else 'N/A'],
            ['Status:', inspection.status.value.upper()],
            ['Created Date:', inspection.created_at.strftime('%Y-%m-%d %H:%M:%S')],
            ['Updated Date:', inspection.updated_at.strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        if inspection.reviewed_by:
            reviewer = db.query(User).filter(User.id == inspection.reviewed_by).first()
            info_data.append(['Reviewed By:', reviewer.username if reviewer else 'N/A'])
            
        if inspection.reviewed_at:
            info_data.append(['Reviewed Date:', inspection.reviewed_at.strftime('%Y-%m-%d %H:%M:%S')])
            
        if inspection.rejection_reason:
            info_data.append(['Rejection Reason:', inspection.rejection_reason])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 30))
        
        # Responses Section
        if responses:
            story.append(Paragraph("RESPONSES", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Get form fields for context
            form_fields = db.query(FormField).filter(FormField.form_id == form.id).all()
            field_map = {field.id: field for field in form_fields}
            
            response_data = [['Question', 'Response', 'Status']]
            
            for response in responses:
                field = field_map.get(response.field_id)
                question = field.field_name if field else f"Field {response.field_id}"
                
                # Format response value
                response_value = ""
                if response.response_value:
                    if field and field.field_type == 'signature':
                        response_value = "[Digital Signature]"
                    elif field and field.field_type == 'photo':
                        response_value = f"[Photo: {response.response_value}]"
                    else:
                        response_value = str(response.response_value)
                
                if response.measurement_value is not None:
                    response_value = str(response.measurement_value)
                    if field and field.field_options and 'unit' in field.field_options:
                        response_value += f" {field.field_options['unit']}"
                
                status = response.pass_hold_status.value.upper() if response.pass_hold_status else ""
                
                response_data.append([question, response_value, status])
            
            response_table = Table(response_data, colWidths=[2.5*inch, 2.5*inch, 1*inch])
            response_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            
            story.append(response_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        print(f"✅ PDF generated: {pdf_path}")
        return pdf_path
        
    except Exception as e:
        print(f"❌ PDF generation failed: {str(e)}")
        return False
    finally:
        db.close()

def generate_sample_excel():
    """Generate a sample Excel file to test formatting"""
    print("Generating sample Excel...")
    
    db = next(get_db())
    
    try:
        # Get sample inspections
        inspections = db.query(Inspection).limit(10).all()
        
        if not inspections:
            print("❌ No inspections found")
            return False
            
        print(f"✓ Using {len(inspections)} inspections")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspections"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get all unique form fields from the inspections
        form_fields_map = {}
        for inspection in inspections:
            form = db.query(Form).filter(Form.id == inspection.form_id).first()
            if form and form.id not in form_fields_map:
                form_fields_map[form.id] = {
                    'name': form.form_name,
                    'fields': sorted(form.fields, key=lambda f: f.field_order)
                }
        
        # Create header row
        headers = [
            "Inspection ID",
            "Form Name",
            "Inspector",
            "Status",
            "Created Date",
            "Updated Date",
            "Reviewed By",
            "Reviewed Date",
            "Rejection Reason"
        ]
        
        # Add dynamic field headers
        all_fields = []
        field_ids_seen = set()
        for form_data in form_fields_map.values():
            for field in form_data['fields']:
                if field.id not in field_ids_seen:
                    all_fields.append(field)
                    field_ids_seen.add(field.id)
                    headers.append(f"{field.field_name} ({field.field_type})")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data rows
        row_num = 2
        for inspection in inspections:
            form = db.query(Form).filter(Form.id == inspection.form_id).first()
            inspector = db.query(User).filter(User.id == inspection.inspector_id).first()
            reviewer = None
            if inspection.reviewed_by:
                reviewer = db.query(User).filter(User.id == inspection.reviewed_by).first()
            
            # Basic inspection info
            row_data = [
                inspection.id,
                form.form_name if form else "N/A",
                inspector.username if inspector else "N/A",
                inspection.status.value.upper(),
                inspection.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                inspection.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                reviewer.username if reviewer else "",
                inspection.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if inspection.reviewed_at else "",
                inspection.rejection_reason or ""
            ]
            
            # Get responses mapped by field_id
            responses_map = {resp.field_id: resp for resp in inspection.responses}
            
            # Add field responses
            for field in all_fields:
                field_response = responses_map.get(field.id)
                cell_value = ""
                
                if field_response:
                    if field_response.response_value:
                        if field.field_type == 'signature':
                            cell_value = "[Digital Signature]"
                        elif field.field_type == 'photo':
                            cell_value = f"[Photo: {field_response.response_value}]"
                        else:
                            cell_value = str(field_response.response_value)
                    
                    if field_response.measurement_value is not None:
                        cell_value = str(field_response.measurement_value)
                        if field.field_options and 'unit' in field.field_options:
                            cell_value += f" {field.field_options['unit']}"
                    
                    if field_response.pass_hold_status:
                        status_text = field_response.pass_hold_status.value.upper()
                        cell_value += f" [{status_text}]" if cell_value else f"[{status_text}]"
                else:
                    cell_value = "—"
                
                row_data.append(cell_value)
            
            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = border
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to file
        excel_filename = f"sample_inspections_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = os.path.join("uploads", excel_filename)
        
        wb.save(excel_path)
        
        print(f"✅ Excel generated: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Excel generation failed: {str(e)}")
        return False
    finally:
        db.close()

if __name__ == "__main__":
    print("🔍 Testing Export Formatting\n")
    
    # Generate sample files
    pdf_path = generate_sample_pdf()
    excel_path = generate_sample_excel()
    
    print(f"\n📊 Generated Files:")
    if pdf_path:
        print(f"PDF: {pdf_path}")
    if excel_path:
        print(f"Excel: {excel_path}")
    
    if pdf_path and excel_path:
        print("\n🎉 Both export formats generated successfully!")
        print("📝 Please review the generated files to verify formatting quality.")
    else:
        print("\n⚠️  Some files failed to generate")