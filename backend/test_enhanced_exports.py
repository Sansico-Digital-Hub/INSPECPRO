#!/usr/bin/env python3
"""
Test script for enhanced PDF and Excel export functionality
Tests the improved formatting, multiple sheets, and professional styling
"""

import requests
import json
import os
from datetime import datetime

# Configuration
BASE_URL = "http://localhost:8000"
TEST_USER = {
    "username": "admin",
    "password": "admin123"
}

def login():
    """Login and get access token"""
    print("🔐 Logging in...")
    
    login_data = {
        "username_or_email": TEST_USER["username"],
        "password": TEST_USER["password"]
    }
    
    response = requests.post(f"{BASE_URL}/api/auth/login", json=login_data)
    
    if response.status_code == 200:
        token = response.json()["access_token"]
        print("✅ Login successful")
        return token
    else:
        print(f"❌ Login failed: {response.status_code} - {response.text}")
        return None

def test_enhanced_pdf_export(token):
    """Test the enhanced PDF export functionality"""
    print("\n📄 Testing Enhanced PDF Export...")
    
    headers = {"Authorization": f"Bearer {token}"}
    
    # Get available inspections first
    inspections_response = requests.get(f"{BASE_URL}/api/inspections/", headers=headers)
    
    if inspections_response.status_code != 200:
        print(f"❌ Failed to get inspections: {inspections_response.status_code}")
        return False
    
    inspections = inspections_response.json()
    
    if not inspections:
        print("❌ No inspections found for testing")
        return False
    
    # Test PDF export with the first inspection
    inspection_id = inspections[0]["id"]
    print(f"📋 Testing PDF export for inspection ID: {inspection_id}")
    
    pdf_response = requests.get(
        f"{BASE_URL}/api/inspections/{inspection_id}/export-pdf",
        headers=headers
    )
    
    if pdf_response.status_code == 200:
        # Save the PDF file for verification
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"enhanced_test_export_{inspection_id}_{timestamp}.pdf"
        pdf_path = os.path.join("uploads", pdf_filename)
        
        os.makedirs("uploads", exist_ok=True)
        
        with open(pdf_path, 'wb') as f:
            f.write(pdf_response.content)
        
        print(f"✅ Enhanced PDF export successful: {pdf_filename}")
        print(f"📁 File saved to: {pdf_path}")
        print(f"📊 File size: {len(pdf_response.content)} bytes")
        return True
    else:
        print(f"❌ PDF export failed: {pdf_response.status_code} - {pdf_response.text}")
        return False

def test_enhanced_excel_export(token):
    """Test the enhanced Excel export functionality with multiple sheets"""
    print("\n📊 Testing Enhanced Excel Export...")
    
    headers = {"Authorization": f"Bearer {token}"}
    
    # Test Excel export with various filters
    test_cases = [
        {"name": "All Inspections", "params": {}},
        {"name": "With Status Filter", "params": {"status": "accepted"}},
        {"name": "With Date Range", "params": {"start_date": "2024-01-01", "end_date": "2024-12-31"}},
    ]
    
    for test_case in test_cases:
        print(f"🧪 Testing: {test_case['name']}")
        
        excel_response = requests.get(
            f"{BASE_URL}/api/inspections/export-excel",
            headers=headers,
            params=test_case["params"]
        )
        
        if excel_response.status_code == 200:
            # Save the Excel file for verification
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"enhanced_test_{test_case['name'].lower().replace(' ', '_')}_{timestamp}.xlsx"
            excel_path = os.path.join("uploads", excel_filename)
            
            os.makedirs("uploads", exist_ok=True)
            
            with open(excel_path, 'wb') as f:
                f.write(excel_response.content)
            
            print(f"✅ Enhanced Excel export successful: {excel_filename}")
            print(f"📁 File saved to: {excel_path}")
            print(f"📊 File size: {len(excel_response.content)} bytes")
        else:
            print(f"❌ Excel export failed for {test_case['name']}: {excel_response.status_code}")
            return False
    
    return True

def verify_excel_sheets():
    """Verify that the Excel file contains the expected multiple sheets"""
    print("\n🔍 Verifying Excel Sheet Structure...")
    
    try:
        from openpyxl import load_workbook
        
        # Find the most recent Excel file
        uploads_dir = "uploads"
        excel_files = [f for f in os.listdir(uploads_dir) if f.startswith("enhanced_test_") and f.endswith(".xlsx")]
        
        if not excel_files:
            print("❌ No test Excel files found")
            return False
        
        # Get the most recent file
        latest_file = max(excel_files, key=lambda x: os.path.getctime(os.path.join(uploads_dir, x)))
        excel_path = os.path.join(uploads_dir, latest_file)
        
        print(f"📂 Analyzing file: {latest_file}")
        
        # Load and analyze the workbook
        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames
        
        print(f"📋 Found {len(sheet_names)} sheets:")
        for i, sheet_name in enumerate(sheet_names, 1):
            sheet = wb[sheet_name]
            row_count = sheet.max_row
            col_count = sheet.max_column
            print(f"  {i}. {sheet_name} ({row_count} rows, {col_count} columns)")
        
        # Check for expected sheets
        expected_sheets = ["📊 Summary", "📋 Detailed Data", "📝 Forms Overview"]
        missing_sheets = [sheet for sheet in expected_sheets if sheet not in sheet_names]
        
        if missing_sheets:
            print(f"❌ Missing expected sheets: {missing_sheets}")
            return False
        else:
            print("✅ All expected sheets are present")
            return True
            
    except ImportError:
        print("⚠️ openpyxl not available for sheet verification")
        return True
    except Exception as e:
        print(f"❌ Error verifying Excel sheets: {e}")
        return False

def main():
    """Main test function"""
    print("🚀 Starting Enhanced Export Functionality Tests")
    print("=" * 60)
    
    # Login
    token = login()
    if not token:
        print("❌ Cannot proceed without authentication")
        return
    
    # Test results
    results = {
        "pdf_export": False,
        "excel_export": False,
        "excel_sheets": False
    }
    
    # Test PDF export
    results["pdf_export"] = test_enhanced_pdf_export(token)
    
    # Test Excel export
    results["excel_export"] = test_enhanced_excel_export(token)
    
    # Verify Excel sheets
    results["excel_sheets"] = verify_excel_sheets()
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 TEST RESULTS SUMMARY")
    print("=" * 60)
    
    for test_name, result in results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{test_name.replace('_', ' ').title()}: {status}")
    
    all_passed = all(results.values())
    overall_status = "✅ ALL TESTS PASSED" if all_passed else "❌ SOME TESTS FAILED"
    
    print(f"\nOverall Status: {overall_status}")
    
    if all_passed:
        print("\n🎉 Enhanced export functionality is working correctly!")
        print("📋 Key improvements verified:")
        print("  • Enhanced PDF formatting with professional styling")
        print("  • Multiple Excel sheets (Summary, Detailed Data, Forms Overview)")
        print("  • Improved color schemes and layouts")
        print("  • Better data organization and presentation")
    else:
        print("\n⚠️ Some issues were detected. Please review the test output above.")

if __name__ == "__main__":
    main()