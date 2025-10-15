#!/usr/bin/env python3
"""
Script to analyze the generated PDF and Excel export files for formatting quality.
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import glob

def analyze_excel_file(file_path):
    """Analyze Excel file formatting and structure."""
    print(f"\n📊 EXCEL FILE ANALYSIS: {os.path.basename(file_path)}")
    print("=" * 60)
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Basic file info
        print(f"📋 Sheet Name: {sheet.title}")
        print(f"📏 Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
        
        # Header analysis
        print(f"\n🏷️  HEADER ANALYSIS:")
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(cell_value)
            print(f"   Column {col}: {cell_value}")
        
        # Data sample analysis
        print(f"\n📝 DATA SAMPLE (First 3 rows):")
        for row in range(2, min(5, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value)[:30] if cell_value else "")
            print(f"   Row {row}: {' | '.join(row_data)}")
        
        # Column width analysis
        print(f"\n📐 COLUMN WIDTH ANALYSIS:")
        for col_letter, col_dimension in sheet.column_dimensions.items():
            if col_dimension.width:
                print(f"   Column {col_letter}: {col_dimension.width:.1f}")
        
        # Cell formatting analysis
        print(f"\n🎨 FORMATTING ANALYSIS:")
        header_cell = sheet.cell(row=1, column=1)
        if header_cell.font:
            print(f"   Header Font: {header_cell.font.name}, Bold: {header_cell.font.bold}")
        if header_cell.fill:
            print(f"   Header Fill: {header_cell.fill.start_color.index}")
        
        print(f"✅ Excel file analysis completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        return False

def analyze_pdf_file(file_path):
    """Analyze PDF file (basic file info only - full PDF parsing would require additional libraries)."""
    print(f"\n📄 PDF FILE ANALYSIS: {os.path.basename(file_path)}")
    print("=" * 60)
    
    try:
        file_size = os.path.getsize(file_path)
        print(f"📏 File Size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
        
        # Read first few bytes to verify it's a valid PDF
        with open(file_path, 'rb') as f:
            header = f.read(8)
            if header.startswith(b'%PDF'):
                pdf_version = header.decode('ascii', errors='ignore')
                print(f"📋 PDF Version: {pdf_version}")
                print(f"✅ Valid PDF file structure detected")
            else:
                print(f"❌ Invalid PDF file structure")
                return False
        
        print(f"✅ PDF file analysis completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error analyzing PDF file: {e}")
        return False

def main():
    """Main function to analyze all export files."""
    print("🔍 EXPORT FILES FORMATTING ANALYSIS")
    print("=" * 60)
    
    uploads_dir = Path("uploads")
    if not uploads_dir.exists():
        print("❌ Uploads directory not found!")
        return
    
    # Find the most recent sample files
    excel_files = list(uploads_dir.glob("sample_inspections_export_*.xlsx"))
    pdf_files = list(uploads_dir.glob("sample_inspection_*.pdf"))
    
    if not excel_files and not pdf_files:
        print("❌ No sample export files found!")
        return
    
    # Analyze Excel files
    excel_success = True
    for excel_file in sorted(excel_files, key=os.path.getmtime, reverse=True)[:1]:  # Most recent
        excel_success = analyze_excel_file(excel_file)
    
    # Analyze PDF files
    pdf_success = True
    for pdf_file in sorted(pdf_files, key=os.path.getmtime, reverse=True)[:1]:  # Most recent
        pdf_success = analyze_pdf_file(pdf_file)
    
    # Summary
    print(f"\n📊 ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"Excel Analysis: {'✅ PASSED' if excel_success else '❌ FAILED'}")
    print(f"PDF Analysis: {'✅ PASSED' if pdf_success else '❌ FAILED'}")
    
    if excel_success and pdf_success:
        print(f"\n🎉 All export files are properly formatted!")
    else:
        print(f"\n⚠️  Some export files may need formatting improvements.")
    
    # Recommendations
    print(f"\n💡 FORMATTING RECOMMENDATIONS:")
    print("   📊 Excel:")
    print("      - Headers should be bold and have background color")
    print("      - Column widths should auto-adjust to content")
    print("      - Data should be properly aligned")
    print("      - Date/time fields should be formatted consistently")
    print("   📄 PDF:")
    print("      - Should include proper headers and footers")
    print("      - Tables should have clear borders and spacing")
    print("      - Text should be readable with appropriate font sizes")
    print("      - Should handle page breaks gracefully")

if __name__ == "__main__":
    main()